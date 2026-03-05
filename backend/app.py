from flask import Flask, request, jsonify
from flask_cors import CORS
import json
import uuid
from datetime import datetime
import os
from zoneinfo import ZoneInfo

from supabase import create_client
from dotenv import load_dotenv

# ✅ SendGrid Imports
from sendgrid import SendGridAPIClient
from sendgrid.helpers.mail import Mail

# ✅ Excel Imports
from openpyxl import Workbook, load_workbook

app = Flask(__name__)
CORS(app)

DATA_FILE = "complaints.json"
EXCEL_FILE = "complaints.xlsx"

# ✅ Load .env
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
ENV_PATH = os.path.join(BASE_DIR, ".env")
load_dotenv(dotenv_path=ENV_PATH)

# ---------------- LOGIN CREDENTIALS ---------------- #
FLAT_CREDENTIALS = {

"A-001": "1234",  "A-002": "1234",  "A-003": "1234",  "A-004": "1234",  "A-005": "1234",
"A-006": "1234",  "A-007": "1234",  "A-008": "1234",  "A-009": "1234",
"A-101": "1234",  "A-102": "1234",  "A-103": "1234",  "A-104": "1234",  "A-105": "1234",
"A-106": "1234",  "A-107": "1234",  "A-108": "1234",  "A-109": "1234",
"A-201": "1234",  "A-202": "1234",  "A-203": "1234",  "A-204": "1234",  "A-205": "1234",
"A-206": "1234",  "A-207": "1234",  "A-208": "1234",  "A-209": "1234",
"A-301": "1234",  "A-302": "1234",  "A-303": "1234",  "A-304": "1234",  "A-305": "1234",
"A-306": "1234",  "A-307": "1234",  "A-308": "1234",  "A-309": "1234",
"A-401": "1234",  "A-402": "1234",  "A-403": "1234",  "A-404": "1234",  "A-405": "1234",
"A-406": "1234",  "A-407": "1234",  "A-408": "1234",  "A-409": "1234",

"B-001": "1234",  "B-002": "1234",  "B-003": "1234",  "B-004": "1234",  "B-005": "1234",
"B-006": "1234",  "B-007": "1234",  "B-008": "1234",  "B-009": "1234",
"B-101": "1234",  "B-102": "1234",  "B-103": "1234",  "B-104": "1234",  "B-105": "1234",
"B-106": "1234",  "B-107": "1234",  "B-108": "1234",  "B-109": "1234",
"B-201": "1234",  "B-202": "1234",  "B-203": "1234",  "B-204": "1234",  "B-205": "1234",
"B-206": "1234",  "B-207": "1234",  "B-208": "1234",  "B-209": "1234",
"B-301": "1234",  "B-302": "1234",  "B-303": "1234",  "B-304": "1234",  "B-305": "1234",
"B-306": "1234",  "B-307": "1234",  "B-308": "1234",  "B-309": "1234",
"B-401": "1234",  "B-402": "1234",  "B-403": "1234",  "B-404": "1234",  "B-405": "1234",
"B-406": "1234",  "B-407": "1234",  "B-408": "1234",  "B-409": "1234",

"C-001": "1234",  "C-002": "1234",  "C-003": "1234",  "C-004": "1234",  "C-005": "1234",
"C-006": "1234",  "C-007": "1234",  "C-008": "1234",  "C-009": "1234",
"C-101": "1234",  "C-102": "1234",  "C-103": "1234",  "C-104": "1234",  "C-105": "1234",
"C-106": "1234",  "C-107": "1234",  "C-108": "1234",  "C-109": "1234",
"C-201": "1234",  "C-202": "1234",  "C-203": "1234",  "C-204": "1234",  "C-205": "1234",
"C-206": "1234",  "C-207": "1234",  "C-208": "1234",  "C-209": "1234",
"C-301": "1234",  "C-302": "1234",  "C-303": "1234",  "C-304": "1234",  "C-305": "1234",
"C-306": "1234",  "C-307": "1234",  "C-308": "1234",  "C-309": "1234",
"C-401": "1234",  "C-402": "1234",  "C-403": "1234",  "C-404": "1234",  "C-405": "1234",
"C-406": "1234",  "C-407": "1234",  "C-408": "1234",  "C-409": "1234",

"D-001": "1234",  "D-002": "1234",  "D-003": "1234",  "D-004": "1234",  "D-005": "1234",
"D-006": "1234",  "D-007": "1234",  "D-008": "1234",  "D-009": "1234",
"D-101": "1234",  "D-102": "1234",  "D-103": "1234",  "D-104": "1234",  "D-105": "1234",
"D-106": "1234",  "D-107": "1234",  "D-108": "1234",  "D-109": "1234",
"D-201": "1234",  "D-202": "1234",  "D-203": "1234",  "D-204": "1234",  "D-205": "1234",
"D-206": "1234",  "D-207": "1234",  "D-208": "1234",  "D-209": "1234",
"D-301": "1234",  "D-302": "1234",  "D-303": "1234",  "D-304": "1234",  "D-305": "1234",
"D-306": "1234",  "D-307": "1234",  "D-308": "1234",  "D-309": "1234",
"D-401": "1234",  "D-402": "1234",  "D-403": "1234",  "D-404": "1234",  "D-405": "1234",
"D-406": "1234",  "D-407": "1234",  "D-408": "1234",  "D-409": "1234",

"E-001": "1234",  "E-002": "1234",  "E-003": "1234",  "E-004": "1234",  "E-005": "1234",
"E-006": "1234",  "E-007": "1234",  "E-008": "1234",  "E-009": "1234",
"E-101": "1234",  "E-102": "1234",  "E-103": "1234",  "E-104": "1234",  "E-105": "1234",
"E-106": "1234",  "E-107": "1234",  "E-108": "1234",  "E-109": "1234",
"E-201": "1234",  "E-202": "1234",  "E-203": "1234",  "E-204": "1234",  "E-205": "1234",
"E-206": "1234",  "E-207": "1234",  "E-208": "1234",  "E-209": "1234",
"E-301": "1234",  "E-302": "1234",  "E-303": "1234",  "E-304": "1234",  "E-305": "1234",
"E-306": "1234",  "E-307": "1234",  "E-308": "1234",  "E-309": "1234",
"E-401": "1234",  "E-402": "1234",  "E-403": "1234",  "E-404": "1234",  "E-405": "1234",
"E-406": "1234",  "E-407": "1234",  "E-408": "1234",  "E-409": "1234",

"F-001": "1234",  "F-002": "1234",  "F-003": "1234",  "F-004": "1234",  "F-005": "1234",
"F-006": "1234",  "F-007": "1234",  "F-008": "1234",  "F-009": "1234",
"F-101": "1234",  "F-102": "1234",  "F-103": "1234",  "F-104": "1234",  "F-105": "1234",
"F-106": "1234",  "F-107": "1234",  "F-108": "1234",  "F-109": "1234",
"F-201": "1234",  "F-202": "1234",  "F-203": "1234",  "F-204": "1234",  "F-205": "1234",
"F-206": "1234",  "F-207": "1234",  "F-208": "1234",  "F-209": "1234",
"F-301": "1234",  "F-302": "1234",  "F-303": "1234",  "F-304": "1234",  "F-305": "1234",
"F-306": "1234",  "F-307": "1234",  "F-308": "1234",  "F-309": "1234",
"F-401": "1234",  "F-402": "1234",  "F-403": "1234",  "F-404": "1234",  "F-405": "1234",
"F-406": "1234",  "F-407": "1234",  "F-408": "1234",  "F-409": "1234"

}


# ---------------- LOGIN API ---------------- #

@app.route("/login", methods=["POST"])
def login():
    data = request.get_json()

    flat = data.get("flatNumber")
    pin = data.get("pin")

    if not flat or not pin:
        return jsonify({"success": False, "message": "Missing credentials"}), 400

    flat = flat.strip().upper()

    if flat in FLAT_CREDENTIALS and FLAT_CREDENTIALS[flat] == pin:
        return jsonify({
            "success": True,
            "message": "Login successful",
            "flatNumber": flat
        })
    else:
        return jsonify({
            "success": False,
            "message": "Invalid Flat Number or PIN"
        }), 401


# -------- Supabase Config --------
SUPABASE_URL = os.getenv("SUPABASE_URL")
SUPABASE_SERVICE_ROLE_KEY = os.getenv("SUPABASE_SERVICE_ROLE_KEY")

supabase = None
if SUPABASE_URL and SUPABASE_SERVICE_ROLE_KEY:
    try:
        supabase = create_client(SUPABASE_URL, SUPABASE_SERVICE_ROLE_KEY)
        print("✅ Supabase client initialized")
    except Exception as e:
        print("❌ Supabase client init error:", e)
else:
    print("⚠️ Supabase not configured.")

# -------- Email Portfolio Mapping --------
PORTFOLIO_EMAILS = {
    "water": "a22853041@gmail.com",
    "electricity": "efgh28099@gmail.com",
    "housekeeping": "ijkl66708@gmail.com",
    "security": "mnop7875@gmail.com"
}

# -------- Utility Functions --------
def load_complaints():
    if not os.path.exists(DATA_FILE):
        return []
    with open(DATA_FILE, "r") as f:
        return json.load(f)

def save_complaints(data):
    with open(DATA_FILE, "w") as f:
        json.dump(data, f, indent=4)

# -------- Excel Save Function --------
def save_to_excel(complaint):
    try:
        if os.path.exists(EXCEL_FILE):
            workbook = load_workbook(EXCEL_FILE)
            sheet = workbook.active
        else:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Complaints"

            headers = [
                "Complaint ID",
                "Flat Number",
                "Name",
                "Phone",
                "Portfolio",
                "Description",
                "Date"
            ]
            sheet.append(headers)

        sheet.append([
            complaint["complaint_id"],
            complaint["flat_number"],
            complaint["name"],
            complaint["phone"],
            complaint["portfolio"],
            complaint["description"],
            complaint["date"]
        ])

        workbook.save(EXCEL_FILE)
        print("✅ Saved to Excel")

    except Exception as e:
        print("❌ Excel save error:", e)

# -------- Email Function --------
def send_email(complaint):
    receiver = PORTFOLIO_EMAILS.get((complaint.get("portfolio") or "").lower())

    if not receiver:
        print("❌ No email mapped for portfolio")
        return

    message = Mail(
        from_email="tribhuvanroy2006@gmail.com",
        to_emails=receiver,
        subject=f"New {complaint['portfolio'].title()} Complaint",
        html_content=f"""
        <h3>New Complaint Registered</h3>
        <p><b>Complaint ID:</b> {complaint['complaint_id']}</p>
        <p><b>Resident:</b> {complaint['name']}</p>
        <p><b>Flat Number:</b> {complaint['flat_number']}</p>
        <p><b>Phone:</b> {complaint['phone']}</p>
        <p><b>Portfolio:</b> {complaint['portfolio']}</p>
        <p><b>Description:</b> {complaint['description']}</p>
        <p><b>Date:</b> {complaint['date']}</p>
        """
    )

    try:
        sg = SendGridAPIClient(os.getenv("SENDGRID_API_KEY"))
        sg.send(message)
        print(f"✅ Email sent to {receiver}")
    except Exception as e:
        print("❌ SendGrid error:", e)

# -------- Supabase Insert --------
def save_to_supabase(complaint):
    if not supabase:
        print("⚠️ Supabase insert skipped.")
        return

    try:
        payload = {
            "complaint_id": complaint["complaint_id"],
            "flat_number": complaint["flat_number"],
            "name": complaint["name"],
            "phone": complaint["phone"],
            "portfolio": complaint["portfolio"],
            "description": complaint["description"],
        }

        res = supabase.table("complaints").insert(payload).execute()
        print("✅ Saved to Supabase:", res.data)

    except Exception as e:
        print("❌ Supabase insert error:", e)

# -------- Home Route --------
@app.route("/", methods=["GET"])
def home():
    return jsonify({
        "status": "OK",
        "message": "Community Connect Backend is running"
    })

# -------- Submit Complaint API --------
@app.route("/submit-complaint", methods=["POST"])
def submit_complaint():
    data = request.get_json()

    if not data:
        return jsonify({"error": "No data received"}), 400

    required_fields = ["flat_number", "resident_name", "phone_number", "portfolio", "description"]

    for field in required_fields:
        if not data.get(field):
            return jsonify({"error": f"{field} is required"}), 400

    complaint_id = str(uuid.uuid4())[:8]

    complaint = {
        "complaint_id": complaint_id,
        "flat_number": data.get("flat_number"),
        "name": data.get("resident_name"),
        "phone": data.get("phone_number"),
        "portfolio": data.get("portfolio"),
        "description": data.get("description"),
        "date": datetime.now(ZoneInfo("Asia/Kolkata")).strftime("%Y-%m-%d --- %H:%M:%S")
    }

    complaints = load_complaints()
    complaints.append(complaint)
    save_complaints(complaints)

    save_to_supabase(complaint)
    save_to_excel(complaint)
    send_email(complaint)

    return jsonify({
        "message": "Complaint submitted successfully",
        "complaint_id": complaint_id
    }), 200

# -------- Run Server --------
if __name__ == "__main__":
    app.run(debug=True)
